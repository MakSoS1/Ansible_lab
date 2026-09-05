from __future__ import annotations

import argparse
import json
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook


def _sample_rows(ws, limit: int = 8) -> list[list[object]]:
    rows: list[list[object]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx >= limit:
            break
        rows.append(list(row))
    return rows


def inspect_archive(archive: Path) -> dict[str, object]:
    with tempfile.TemporaryDirectory(prefix="aios-chdd-") as tmp:
        root = Path(tmp)
        with zipfile.ZipFile(archive) as zf:
            zf.extractall(root)
        files = sorted(p for p in root.rglob("*") if p.is_file())
        python_files: dict[str, str] = {}
        workbooks: dict[str, object] = {}
        invalid_workbooks: dict[str, str] = {}
        for path in files:
            rel = str(path.relative_to(root))
            if path.suffix.lower() == ".py":
                python_files[rel] = path.read_text(encoding="utf-8", errors="replace")
            elif path.suffix.lower() in {".xlsx", ".xlsm"}:
                if path.name.startswith("~$"):
                    invalid_workbooks[rel] = "temporary Excel lock file"
                    continue
                try:
                    wb = load_workbook(path, read_only=True, data_only=True)
                except (zipfile.BadZipFile, OSError, ValueError) as exc:
                    invalid_workbooks[rel] = f"{type(exc).__name__}: {exc}"
                    continue
                workbooks[rel] = {
                    "sheets": [
                        {
                            "title": ws.title,
                            "max_row": ws.max_row,
                            "max_column": ws.max_column,
                            "sample_rows": _sample_rows(ws),
                        }
                        for ws in wb.worksheets
                    ]
                }
                wb.close()
        return {
            "archive": str(archive),
            "files": [str(p.relative_to(root)) for p in files],
            "python_files": python_files,
            "workbooks": workbooks,
            "invalid_workbooks": invalid_workbooks,
        }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("archive", type=Path)
    args = parser.parse_args()
    print(json.dumps(inspect_archive(args.archive), ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()
