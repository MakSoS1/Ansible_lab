from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class ReferenceParityThresholds:
    max_npv_relative_error: float = 0.005
    max_mean_physical_relative_error: float = 0.01
    max_annual_physical_relative_error: float = 0.02
    max_mean_active_well_month_relative_error: float = 0.02


DEFAULT_REFERENCE_PARITY_THRESHOLDS = ReferenceParityThresholds()

_REFERENCE_ROWS = {
    "oil_kt": "Добыча нефти",
    "liquid_kt": "Добыча жидкости",
    "injection_km3": "Закачка",
    "active_well_months": "Средний действующий фонд",
    "fcf_mrub": "FCF",
    "chdd_mrub": "ЧДД",
}


def _finite_number(value: Any) -> float:
    result = float(value or 0.0)
    if not np.isfinite(result):
        raise ValueError(f"reference workbook contains non-finite value: {value!r}")
    return result


def load_model_z_reference(workbook_path: Path, *, start_year: int = 2007, wacc: float = 0.10) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError(f"expected one organizer worksheet, found {workbook.sheetnames}")
    sheet = workbook[workbook.sheetnames[0]]
    years: list[int] = []
    columns: list[int] = []
    for column in range(3, sheet.max_column + 1):
        value = sheet.cell(1, column).value
        if value is None:
            continue
        try:
            year = int(str(value))
        except ValueError as exc:
            raise ValueError(f"unexpected year header {value!r}") from exc
        if year >= start_year:
            years.append(year)
            columns.append(column)
    if not years:
        raise ValueError(f"reference workbook has no years >= {start_year}")

    row_by_name = {
        str(sheet.cell(row, 1).value).strip(): row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value is not None
    }
    missing = sorted(label for label in _REFERENCE_ROWS.values() if label not in row_by_name)
    if missing:
        raise ValueError(f"reference workbook is missing rows: {missing}")

    annual: list[dict[str, float | int]] = []
    for year, column in zip(years, columns, strict=True):
        annual.append(
            {
                "year": year,
                **{
                    key: _finite_number(sheet.cell(row_by_name[label], column).value)
                    for key, label in _REFERENCE_ROWS.items()
                },
            }
        )
    rebased_npv = sum(
        float(row["fcf_mrub"]) / ((1.0 + wacc) ** (int(row["year"]) - start_year))
        for row in annual
    )
    original_discounted_sum = sum(float(row["chdd_mrub"]) for row in annual)
    return {
        "sheet": sheet.title,
        "start_year": start_year,
        "wacc": float(wacc),
        "annual": annual,
        "rebased_npv_mrub": float(rebased_npv),
        "original_workbook_chdd_sum_mrub": float(original_discounted_sum),
    }


def _relative_error(observed: float, reference: float) -> float:
    return float(abs(observed - reference) / max(abs(reference), 1e-12))


def reference_parity_report(
    calculation: dict[str, Any],
    reference: dict[str, Any],
    *,
    thresholds: ReferenceParityThresholds = DEFAULT_REFERENCE_PARITY_THRESHOLDS,
) -> dict[str, Any]:
    calculated_by_year = {int(row["year"]): row for row in calculation["annual"]}
    reference_by_year = {int(row["year"]): row for row in reference["annual"]}
    common_years = sorted(set(calculated_by_year) & set(reference_by_year))
    if common_years != sorted(reference_by_year):
        missing = sorted(set(reference_by_year) - set(calculated_by_year))
        raise ValueError(f"calculation is missing organizer reference years: {missing}")

    field_pairs = {
        "oil": ("oilKt", "oil_kt"),
        "liquid": ("liquidKt", "liquid_kt"),
        "injection": ("injectionKm3", "injection_km3"),
    }
    annual_physical: dict[str, list[float]] = {name: [] for name in field_pairs}
    active_errors: list[float] = []
    annual_rows: list[dict[str, Any]] = []
    for year in common_years:
        calculated = calculated_by_year[year]
        expected = reference_by_year[year]
        row_errors: dict[str, float] = {}
        for name, (calc_key, ref_key) in field_pairs.items():
            error = _relative_error(float(calculated[calc_key]), float(expected[ref_key]))
            annual_physical[name].append(error)
            row_errors[f"{name}_relative_error"] = error
        active_error = _relative_error(float(calculated["activeWellMonths"]), float(expected["active_well_months"]))
        active_errors.append(active_error)
        row_errors["active_well_month_relative_error"] = active_error
        annual_rows.append({"year": year, **row_errors})

    all_physical_errors = [error for values in annual_physical.values() for error in values]
    calculated_npv = float(calculation["summary"]["totalChddM"])
    reference_npv = float(reference["rebased_npv_mrub"])
    npv_error = _relative_error(calculated_npv, reference_npv)
    mean_physical = float(np.mean(all_physical_errors))
    max_physical = float(np.max(all_physical_errors))
    mean_active = float(np.mean(active_errors))
    failures: list[str] = []
    if npv_error > thresholds.max_npv_relative_error:
        failures.append("NPV_PARITY")
    if mean_physical > thresholds.max_mean_physical_relative_error:
        failures.append("MEAN_PHYSICAL_PARITY")
    if max_physical > thresholds.max_annual_physical_relative_error:
        failures.append("ANNUAL_PHYSICAL_PARITY")
    if mean_active > thresholds.max_mean_active_well_month_relative_error:
        failures.append("ACTIVE_FUND_PARITY")
    return {
        "passed": not failures,
        "failures": failures,
        "calculated_npv_mrub": calculated_npv,
        "reference_npv_mrub": reference_npv,
        "npv_relative_error": npv_error,
        "mean_physical_relative_error": mean_physical,
        "max_annual_physical_relative_error": max_physical,
        "mean_active_well_month_relative_error": mean_active,
        "annual": annual_rows,
        "thresholds": {
            "max_npv_relative_error": thresholds.max_npv_relative_error,
            "max_mean_physical_relative_error": thresholds.max_mean_physical_relative_error,
            "max_annual_physical_relative_error": thresholds.max_annual_physical_relative_error,
            "max_mean_active_well_month_relative_error": thresholds.max_mean_active_well_month_relative_error,
        },
    }
