from pathlib import Path

import pytest
from openpyxl import Workbook

from aios_track2.reference_parity import load_model_z_reference, reference_parity_report


def _calculation() -> dict:
    return {
        "summary": {"totalChddM": 190.9090909090909},
        "annual": [
            {"year": 2007, "oilKt": 100.0, "liquidKt": 200.0, "injectionKm3": 300.0, "activeWellMonths": 600.0},
            {"year": 2008, "oilKt": 90.0, "liquidKt": 210.0, "injectionKm3": 310.0, "activeWellMonths": 590.0},
        ],
    }


def _reference() -> dict:
    return {
        "rebased_npv_mrub": 190.9090909090909,
        "annual": [
            {"year": 2007, "oil_kt": 100.0, "liquid_kt": 200.0, "injection_km3": 300.0, "active_well_months": 600.0},
            {"year": 2008, "oil_kt": 90.0, "liquid_kt": 210.0, "injection_km3": 310.0, "active_well_months": 590.0},
        ],
    }


def test_reference_parity_passes_exact_match_and_fails_material_drift() -> None:
    exact = reference_parity_report(_calculation(), _reference())
    assert exact["passed"] is True
    assert exact["npv_relative_error"] == pytest.approx(0.0)

    drifted = _calculation()
    drifted["summary"] = {"totalChddM": 180.0}
    drifted["annual"] = [dict(row) for row in drifted["annual"]]
    drifted["annual"][1]["oilKt"] = 80.0
    report = reference_parity_report(drifted, _reference())
    assert report["passed"] is False
    assert "NPV_PARITY" in report["failures"]
    assert "ANNUAL_PHYSICAL_PARITY" in report["failures"]


def test_reference_workbook_rebases_fcf_to_track2_start(tmp_path: Path) -> None:
    path = tmp_path / "reference.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Model Z"
    sheet.cell(1, 3, 2006)
    sheet.cell(1, 4, 2007)
    sheet.cell(1, 5, 2008)
    rows = {
        "Добыча нефти": (1.0, 100.0, 90.0),
        "Добыча жидкости": (2.0, 200.0, 210.0),
        "Закачка": (3.0, 300.0, 310.0),
        "Средний действующий фонд": (4.0, 600.0, 590.0),
        "FCF": (5.0, 100.0, 100.0),
        "ЧДД": (6.0, 50.0, 40.0),
    }
    for row_index, (label, values) in enumerate(rows.items(), start=2):
        sheet.cell(row_index, 1, label)
        for column, value in zip((3, 4, 5), values, strict=True):
            sheet.cell(row_index, column, value)
    workbook.save(path)

    reference = load_model_z_reference(path, start_year=2007, wacc=0.10)
    assert [row["year"] for row in reference["annual"]] == [2007, 2008]
    assert reference["rebased_npv_mrub"] == pytest.approx(100.0 + 100.0 / 1.1)
    assert reference["original_workbook_chdd_sum_mrub"] == pytest.approx(90.0)
